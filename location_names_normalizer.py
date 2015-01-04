#!C:\Python27
# -*- coding: utf-8 -*-
import sys
from openpyxl import load_workbook
from utils import write_ws
from location_lists import LocationsList, LocationsDict
from location_lists import NormalizedLocationsList


class LocationsFile():
    """Parse the location files into location lists and creates normalized
    list between two location lists."""

    # DATA
    INPUT_FILE = "location_names.xlsx"
    OUTPUT_FILE = "location_names_normalized.xlsx"
    NORMALIZED_SHEET = "tbl_normal"
    NOT_FOUND_SHEET = "not_found"

    def __init__(self, wb):
        self.wb = wb
        self.first_list = LocationsList(wb)
        self.second_list = LocationsDict(wb)
        self.normalized_list = NormalizedLocationsList()
        self.not_found_list = []

    # PUBLIC
    def normalize_names(self):
        """Looks in second_list for each location in first_list, when there is
        a match, adds to normalized list."""

        # clear normalized list
        self.normalized_list = NormalizedLocationsList()
        self.not_found_list = []

        # iterate through all first_list locations
        i = 1
        for location in self.first_list:

            # print "Looking number", i, "of", len(self.first_list),
            location_matched = self.second_list.find(location)

            # if there is a match, add to normalized list
            if location_matched:
                self.normalized_list.add(location, location_matched)

            else:
                # self.not_found_list.append(location)
                # print "no match,", location
                pass

            i += 1

    def count_first_list(self):
        return self.first_list.count()

    def count_second_list(self):
        return self.second_list.count()

    def save(self, output_file=None):
        """Add normalized list and not found list to workbook and save it."""

        # take default values
        output_file = output_file or self.OUTPUT_FILE

        # add lists to workbook, creating new sheets
        self._list_to_sheet(self.normalized_list, self.NORMALIZED_SHEET)
        self._list_to_sheet(self.not_found_list, self.NOT_FOUND_SHEET)

        # save excel with normalized list of locations
        self.wb.save(output_file)

    # PRIVATE
    def _list_to_sheet(self, locations_list, sheet_name):

        # create new sheet
        ws = self.wb.create_sheet(-1, sheet_name)

        # write each record in excel sheet
        for record in self.normalized_list:
            write_ws(ws, record)


# DATA
ROOT_NAME = "location_names"
INPUT_FILE = ROOT_NAME + ".xlsx"
OUTPUT_FILE = ROOT_NAME + "_normalized.xlsx"


# USER METHODS
def normalize_location_names(input_file=None, output_file=None):
    """Takes a workbook with a list of locations at first two sheets and
    add a sheet with correspondence table between location names at them
    two lists."""

    # if not i/o files are passed, defaults are used
    input_file = input_file or INPUT_FILE
    output_file = output_file or OUTPUT_FILE

    # loads excel file
    wb = load_workbook(input_file)

    # creates a location names file object
    lf = LocationsFile(wb)

    # count records in each list
    # print "First list has", lf.count_first_list(), "records"
    # print "Second list has", lf.count_second_list(), "records"

    # creates normalized list of names
    lf.normalize_names()

    # count records in normalized list
    # print "Normalized list has", lf.count_normalized_list(), "records"

    # save excel workbook with normalized sheet appended
    lf.save(output_file)


# executes main routine
if __name__ == '__main__':

    # if parameters are passed, use them
    input_file = None
    output_file = None

    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]

    # main call
    # normalize_location_names(input_file, output_file)
    normalize_location_names()
